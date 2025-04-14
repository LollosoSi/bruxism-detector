import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.chart.axis import TextAxis
from openpyxl.chart import ScatterChart

# Load CSV (semicolon-delimited, comma as decimal)
df = pd.read_csv("Summary/summary.csv", sep=';', decimal=',')

# Fix merged column if necessary
merged_col = "Stopped after beep; Did not stop after beeps"
if merged_col in df.columns:
    split_col = df[merged_col].astype(str).str.split(";", expand=True)
    df["Stopped after beep"] = pd.to_numeric(split_col[0], errors='coerce')
    df["Did not stop after beeps"] = pd.to_numeric(split_col[1], errors='coerce')
    df.drop(columns=[merged_col], inplace=True)

# Convert all columns that can be parsed as numbers
df = df.apply(pd.to_numeric, errors='ignore')

# Create workbook and worksheet
wb = Workbook()
ws = wb.active
ws.title = "Clenching Charts"

# Shift data to the right (e.g., start at column AZ)
data_col_offset = 29  # Column AC
data_start_col_letter = get_column_letter(data_col_offset)

# Write headers to Excel
for col_idx, col_name in enumerate(df.columns, data_col_offset):
    ws.cell(row=1, column=col_idx, value=col_name)

# Write data rows to Excel
for row_idx, row in enumerate(df.itertuples(index=False), 2):
    for col_idx, value in enumerate(row, data_col_offset):
        ws.cell(row=row_idx, column=col_idx, value=value)

# Chart definitions (column name, title)
charts = [
    ("Clenching Rate (per hour)", "Clenching Rate"),
    ("Jaw Events", "Jaw Events"),
    ("Alarm Triggers", "Alarm Triggers"),
    ("Did not stop after beeps", "Did Not Stop After Beep"),
    ("Avg beeps per event", "Avg Beeps per Event"),
    ("Average clenching duration (seconds)", "Avg Duration (s)"),
    ("Average clenching event pause (minutes)", "Avg Pause (min)"),
]

# Chart layout parameters
chart_height = 8
chart_width = 15
cols_in_grid = 3
vertical_spacing = 8
horizontal_spacing = -6
chart_start_col = 1
chart_start_row = 1

# Create and place charts
for idx, (col_name, chart_title) in enumerate(charts):
    if col_name not in df.columns:
        print(f"⚠️ Skipping missing column: {col_name}")
        continue

    x_col = df.columns.get_loc("Date") + data_col_offset
    y_col = df.columns.get_loc(col_name) + data_col_offset
    max_row = len(df) + 1

    # Create LineChart
    chart = LineChart()
    chart.title = chart_title
    chart.y_axis.title = chart_title
    chart.x_axis.title = "Date"
    chart.height = chart_height
    chart.width = chart_width
    chart.style = None  # Default Excel chart style

    # Add data
    values = Reference(ws, min_col=y_col, min_row=1, max_row=max_row)
    dates = Reference(ws, min_col=x_col, min_row=2, max_row=max_row)
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(dates)
    chart.y_axis.majorGridlines = None  # (see next section)
    chart.y_axis.scaling.min = None
    chart.y_axis.scaling.max = None
    chart.y_axis.majorGridlines = None
    chart.x_axis.majorGridlines = None


    # Styling: disable smoothing, use default blue line, no markers
    series = chart.series[0]
    series.smooth = False
    series.graphicalProperties.line.solidFill = "4F81BD"  # standard blue
    series.marker = None

    # Rotate date labels vertically
    chart.x_axis.label_rotation = 90
    chart.x_axis.axPos = 'b'  # Place X-axis at the bottom
    chart.x_axis = TextAxis()
    chart.x_axis.title = "Date"
    chart.x_axis.axPos = 'b'
    chart.x_axis.tickLblPos = 'low'           # Places labels low on X-axis
    chart.x_axis.textRotation = -45

    
    # After you create the chart, and after setting chart titles, add:
    if col_name == "Avg beeps per event":
        data_values = df[col_name].dropna()
        if not data_values.empty:
            ymin = data_values.min()
            ymax = data_values.max()
            padding = (ymax - ymin) * 0.2 if ymax != ymin else 1
            chart.y_axis.scaling.min = max(0, ymin - padding)
            chart.y_axis.scaling.max = ymax + padding

    # Position in a grid layout (e.g., 2 columns)
    chart_col = chart_start_col + (idx % cols_in_grid) * (chart_width + horizontal_spacing)
    chart_row = chart_start_row + (idx // cols_in_grid) * (chart_height + vertical_spacing)
    cell = f"{get_column_letter(chart_col)}{chart_row}"
    ws.add_chart(chart, cell)

# Save to file
wb.save("output.xlsx")
